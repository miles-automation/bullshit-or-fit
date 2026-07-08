"""WARN Act filings — a supply-side source (workers entering the market).

Under the federal WARN Act, larger employers must file advance notice of mass
layoffs/closures with their state. States publish that database in different
shapes: Texas + Oregon as Socrata JSON feeds, California as an .xlsx export from
EDD (a rolling current-fiscal-year file, our highest tech-signal source). Each is
a WARN_SOURCES entry with a `fmt` and a per-state parser; more states slot in the
same way.

WARN history is immutable — a past filing doesn't change — so ingestion upserts on
a deterministic id and is fully idempotent (no snapshot/close-detection like the
demand-side boards). All parsers are pure; WarnClient is httpx with an injectable
transport for tests.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

import httpx
from sqlalchemy import delete, func, insert, select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.orm import Session

from app.config import settings
from app.jobtrends.models import WarnMonth, WarnNotice

logger = logging.getLogger(__name__)

SOURCE_WARN = "warn"


@dataclass(frozen=True)
class ParsedWarn:
    id: str
    state: str
    company: str
    city: str | None
    employees_affected: int | None
    notice_date: date | None
    effective_date: date | None
    layoff_type: str | None


def _warn_date(value: Any) -> date | None:
    if not isinstance(value, str) or len(value) < 10:
        return None
    try:
        return date.fromisoformat(value[:10])
    except ValueError:
        return None


def _warn_int(value: Any) -> int | None:
    try:
        n = int(float(str(value).replace(",", "")))
    except (TypeError, ValueError):
        return None
    return n if n >= 0 else None


def _slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", (text or "").lower()).strip("-") or "unknown"


def _warn_id(state: str, external_id: str | None, *parts: Any) -> str:
    """Stable id: the state's own notice number when present, else a deterministic
    composite of the discriminating fields (date, company, city, employees…) —
    enough to dedupe re-fetches of immutable history AND to keep two distinct
    same-day filings by one company from collapsing."""
    if external_id:
        return f"{state}:{external_id}"
    tail = ":".join(_slug(str(p)) for p in parts if p not in (None, "")) or "na"
    return f"{state}:{tail}"


def parse_texas(rows: Any) -> list[ParsedWarn]:
    """data.texas.gov/8w53-c4f6 -> ParsedWarn[]. Pure."""
    out: list[ParsedWarn] = []
    for r in rows or []:
        company = str(r.get("job_site_name") or "").strip()
        if not company:
            continue
        notice = _warn_date(r.get("notice_date"))
        city = r.get("city_name") or None
        effective = _warn_date(r.get("layoff_date"))
        employees = _warn_int(r.get("total_layoff_number"))
        out.append(
            ParsedWarn(
                id=_warn_id("TX", None, notice, company, city, effective, employees),
                state="TX",
                company=company,
                city=city,
                employees_affected=employees,
                notice_date=notice,
                effective_date=effective,
                layoff_type=None,
            )
        )
    return out


def parse_oregon(rows: Any) -> list[ParsedWarn]:
    """data.oregon.gov/ijbz-jpx8 -> ParsedWarn[]. Pure."""
    out: list[ParsedWarn] = []
    for r in rows or []:
        company = str(r.get("company_name") or "").strip()
        if not company:
            continue
        notice = _warn_date(r.get("received_date"))
        # NB: Oregon's `warn` number is NOT row-unique — one filing spans multiple
        # site rows with different `laid_off` counts (e.g. a 4-site notice totaling
        # 669). So each row is its own site-level record; the id is a composite of
        # warn# + the row's discriminators, which keeps distinct sites separate and
        # still dedupes exact duplicate rows.
        out.append(
            ParsedWarn(
                id=_warn_id(
                    "OR",
                    None,
                    r.get("warn"),
                    notice,
                    company,
                    r.get("city"),
                    r.get("laid_off"),
                ),
                # Attribute to the source database (Oregon), not the row's `state`
                # field — that's the employer's HQ and is ~93% OR with stray values
                # that would falsely imply multi-state coverage.
                state="OR",
                company=company,
                city=(r.get("city") or None),
                employees_affected=_warn_int(r.get("laid_off")),
                notice_date=notice,
                effective_date=_warn_date(r.get("layoff_date")),
                layoff_type=(r.get("layoff_type") or None),
            )
        )
    return out


def _excel_date(value: Any) -> date | None:
    """openpyxl (data_only) hands back datetimes; tolerate ISO strings too."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return _warn_date(value)


def parse_california(content: bytes) -> list[ParsedWarn]:
    """California EDD 'Detailed WARN Report' .xlsx bytes -> ParsedWarn[]. Pure.

    Columns: County | Notice Date | Processed Date | Effective Date | Company |
    Layoff/Closure | No. Of Employees | Address | Industry. Rows 1–2 are the title
    and header; data follows. The file is a rolling current-fiscal-year report.
    """
    import io
    import warnings

    import openpyxl

    out: list[ParsedWarn] = []
    # read-only openpyxl warns (lazily, during iteration) about an unsupported
    # data-validation extension — suppress it across the whole read.
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = next(
            (s for s in wb.sheetnames if s.strip().lower().startswith("detailed warn")),
            wb.sheetnames[0],
        )
        for row in wb[sheet].iter_rows(min_row=3, values_only=True):
            cols = list(row) + [None] * (9 - len(row))
            (
                county,
                notice,
                _processed,
                effective,
                company_raw,
                ltype,
                emp,
                address,
                _ind,
            ) = cols[:9]
            company = str(company_raw or "").strip()
            if not company:
                continue
            notice_d = _excel_date(notice)
            county_s = str(county or "").strip() or None
            out.append(
                ParsedWarn(
                    id=_warn_id(
                        "CA", None, notice_d, company, county_s, address, _warn_int(emp)
                    ),
                    state="CA",
                    company=company,
                    city=county_s,  # EDD gives county, not city
                    employees_affected=_warn_int(emp),
                    notice_date=notice_d,
                    effective_date=_excel_date(effective),
                    layoff_type=str(ltype or "").strip() or None,
                )
            )
    return out


@dataclass(frozen=True)
class WarnSource:
    state: str
    url: str
    parser: Any  # Callable[[Any], list[ParsedWarn]]
    fmt: str = "socrata"  # 'socrata' (JSON) | 'excel' (.xlsx bytes)
    order_field: str | None = None


WARN_SOURCES: list[WarnSource] = [
    WarnSource(
        "TX",
        "https://data.texas.gov/resource/8w53-c4f6.json",
        parse_texas,
        order_field="notice_date",
    ),
    WarnSource(
        "OR",
        "https://data.oregon.gov/resource/ijbz-jpx8.json",
        parse_oregon,
        order_field="received_date",
    ),
    WarnSource(
        "CA",
        "https://edd.ca.gov/siteassets/files/jobs_and_training/warn/warn_report1.xlsx",
        parse_california,
        fmt="excel",
    ),
]


class WarnClient:
    def __init__(
        self,
        *,
        timeout_seconds: float = 60.0,
        transport: httpx.BaseTransport | None = None,
        user_agent: str | None = None,
    ) -> None:
        self.timeout_seconds = timeout_seconds
        self._transport = transport
        self._user_agent = user_agent or settings.jobtrends_user_agent

    def fetch(self, source: WarnSource) -> list[ParsedWarn]:
        with httpx.Client(
            timeout=self.timeout_seconds,
            transport=self._transport,
            follow_redirects=True,
        ) as client:
            if source.fmt == "excel":
                resp = client.get(source.url, headers={"User-Agent": self._user_agent})
                resp.raise_for_status()
                return source.parser(resp.content)
            params = {
                "$order": f"{source.order_field} DESC",
                "$limit": str(settings.warn_max_records),
            }
            resp = client.get(
                source.url, params=params, headers={"User-Agent": self._user_agent}
            )
            resp.raise_for_status()
            return source.parser(resp.json())


def _upsert(session: Session, notices: list[ParsedWarn]) -> None:
    if not notices:
        return
    # Dedupe by id within the batch (last wins): Postgres rejects an
    # INSERT…ON CONFLICT whose proposed rows share a conflict key, and a state's
    # feed can list the same notice twice.
    unique = {n.id: n for n in notices}
    rows = [
        {
            "id": n.id,
            "state": n.state,
            "company": n.company,
            "city": n.city,
            "employees_affected": n.employees_affected,
            "notice_date": n.notice_date,
            "effective_date": n.effective_date,
            "layoff_type": n.layoff_type,
        }
        for n in unique.values()
    ]
    stmt = pg_insert(WarnNotice).values(rows)
    stmt = stmt.on_conflict_do_update(
        index_elements=[WarnNotice.id],
        set_={
            "company": stmt.excluded.company,
            "city": stmt.excluded.city,
            "employees_affected": stmt.excluded.employees_affected,
            "notice_date": stmt.excluded.notice_date,
            "effective_date": stmt.excluded.effective_date,
            "layoff_type": stmt.excluded.layoff_type,
        },
    )
    session.execute(stmt)


def warn_ingest(
    session: Session, client: WarnClient, sources: list[WarnSource] | None = None
) -> dict[str, int]:
    """Ingest each state's WARN feed (idempotent upsert). One bad state is logged
    and skipped. Returns {state: rows}."""
    result: dict[str, int] = {}
    for source in sources or WARN_SOURCES:
        try:
            notices = client.fetch(source)
        except Exception:  # noqa: BLE001 — one state must not sink the ingest
            logger.exception("jobtrends: WARN fetch failed for %s", source.state)
            continue
        _upsert(session, notices)
        session.commit()
        result[source.state] = len(notices)
        logger.info("jobtrends: WARN %s — %s notices", source.state, len(notices))
    total = session.scalar(select(func.count()).select_from(WarnNotice))
    logger.info("jobtrends: WARN ingest complete — %s notices stored", total)
    return result


def extract_warn_month(session: Session) -> dict[str, int]:
    """Rebuild `warn_month` from warn_notices, bucketed by notice-date month."""
    counts: dict[date, list[int]] = {}
    for notice_date, employees in session.execute(
        select(WarnNotice.notice_date, WarnNotice.employees_affected).where(
            WarnNotice.notice_date.is_not(None)
        )
    ).yield_per(1000):
        month = notice_date.replace(day=1)
        bucket = counts.setdefault(month, [0, 0])
        bucket[0] += 1
        bucket[1] += int(employees or 0)

    rows = [
        {"month": month, "notices": n, "employees_affected": emp}
        for month, (n, emp) in counts.items()
    ]
    session.execute(delete(WarnMonth))
    if rows:
        session.execute(insert(WarnMonth), rows)
    session.commit()
    logger.info("jobtrends: warn_month rebuilt — %s months", len(rows))
    return {"months": len(rows)}


@dataclass(frozen=True)
class WarnMonthOut:
    month: str  # 'YYYY-MM'
    notices: int
    employees_affected: int


def warn_months(session: Session) -> list[WarnMonthOut]:
    """Per-month WARN filings + employees affected, oldest→newest."""
    rows = session.execute(
        select(
            WarnMonth.month, WarnMonth.notices, WarnMonth.employees_affected
        ).order_by(WarnMonth.month)
    ).all()
    return [
        WarnMonthOut(month=m.strftime("%Y-%m"), notices=n, employees_affected=emp)
        for m, n, emp in rows
    ]


@dataclass(frozen=True)
class WarnNoticeOut:
    company: str
    state: str
    city: str | None
    employees_affected: int | None
    notice_date: str | None


@dataclass(frozen=True)
class StateCount:
    state: str
    notices: int
    employees_affected: int


@dataclass(frozen=True)
class WarnReport:
    total_notices: int
    total_employees: int
    states: list[str]
    recent: list[WarnNoticeOut]
    by_state: list[StateCount]


def warn_report(session: Session, *, recent_limit: int = 10) -> WarnReport:
    """Recent notices + per-state totals across the tracked states."""
    by_state_rows = session.execute(
        select(
            WarnNotice.state,
            func.count(),
            func.coalesce(func.sum(WarnNotice.employees_affected), 0),
        ).group_by(WarnNotice.state)
    ).all()
    by_state = [
        StateCount(state=s, notices=int(n), employees_affected=int(emp))
        for s, n, emp in by_state_rows
    ]
    by_state.sort(key=lambda x: x.employees_affected, reverse=True)

    recent_rows = session.execute(
        select(
            WarnNotice.company,
            WarnNotice.state,
            WarnNotice.city,
            WarnNotice.employees_affected,
            WarnNotice.notice_date,
        )
        .where(WarnNotice.notice_date.is_not(None))
        .order_by(WarnNotice.notice_date.desc())
        .limit(recent_limit)
    ).all()
    recent = [
        WarnNoticeOut(
            company=c,
            state=s,
            city=city,
            employees_affected=emp,
            notice_date=nd.isoformat() if nd else None,
        )
        for c, s, city, emp, nd in recent_rows
    ]

    return WarnReport(
        total_notices=sum(x.notices for x in by_state),
        total_employees=sum(x.employees_affected for x in by_state),
        states=sorted(x.state for x in by_state),
        recent=recent,
        by_state=by_state,
    )


def format_warn_table(report: WarnReport) -> str:
    if not report.total_notices:
        return "no data — run `warn-ingest` first"
    lines = [
        f"{report.total_notices} WARN filings / "
        f"{report.total_employees:,} employees affected "
        f"({', '.join(report.states)})",
        "",
        "recent notices:",
    ]
    for n in report.recent:
        emp = f"{n.employees_affected:,}" if n.employees_affected is not None else "?"
        lines.append(f"  {n.notice_date or '?':<12}{n.state}  {emp:>7}  {n.company}")
    return "\n".join(lines)
